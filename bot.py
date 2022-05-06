import logging
import schedule
import time
import openpyxl
from openpyxl import Workbook



from aiogram import Bot, types
from aiogram.utils import executor
from aiogram.dispatcher import Dispatcher
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.contrib.middlewares.logging import LoggingMiddleware

from config import TOKEN
from utils import TestStates
from messages import MESSAGES

logging.basicConfig(format=u'%(filename)+13s [ LINE:%(lineno)-4s] %(levelname)-8s [%(asctime)s] %(message)s',
                    level=logging.DEBUG)

bot = Bot(token=TOKEN)
dp = Dispatcher(bot, storage=MemoryStorage())
dp.middleware.setup(LoggingMiddleware())

ws = openpyxl.load_workbook('ques1.xlsx')
sh1 = ws['Sheet1']
sh2 = ws['Sheet2']
sh3 = ws['Sheet3']

users = {}


def save_s(message):
    wb = openpyxl.load_workbook('users.xlsx')
    sheet = wb['users']
    sheet = wb.active
    cell = sheet.cell(row=users[message.from_user.id][0] + 2, column=3)
    cell.value = users[message.from_user.id][1]
    wb.save('users.xlsx')
    wb.close()


@dp.message_handler(commands=['start'])
async def process_start_command(message: types.Message):
    state = dp.current_state(user=message.from_user.id)
    await message.answer(MESSAGES['start'])
    users[message.from_user.id] = len(users), 0, 1  # us, score, qs

    wb = openpyxl.load_workbook('users.xlsx')
    sheet = wb['users']
    sheet = wb.active
    cell = sheet.cell(row=users[message.from_user.id][0] + 2, column=1)
    cell.value = message.from_user.id
    cell = sheet.cell(row=users[message.from_user.id][0] + 2, column=3)
    cell.value = 0
    cell = sheet.cell(row=users[message.from_user.id][0] + 2, column=5)
    cell.value = 1
    wb.save('users.xlsx')
    wb.close()
    await state.set_state(TestStates.all()[6])


@dp.message_handler(state='*', commands=['info'])
async def process_help_command(message: types.Message):
    await message.answer(MESSAGES['info'])


@dp.message_handler(state='*', commands=['thanks'])
async def process_thx_command(message: types.Message):
    await message.answer(MESSAGES['thx'])


@dp.message_handler(state=TestStates.TEST_STATE_2 | TestStates.TEST_STATE_0, commands=['quiz'])
async def process_quiz_command(message: types.Message):
    w = openpyxl.load_workbook('users.xlsx')
    s = w['users']
    s = w.active
    if users[message.from_user.id][2] == 1:
        await message.answer(MESSAGES['first_s'])
    if str(s.cell(row=users[message.from_user.id][0] + 2, column=5).value) == str(1) and users[message.from_user.id][2]\
            > 20:
        await message.answer(MESSAGES['second_s'])
        await message.answer('Введите ответ на 1 вопрос')
    elif users[message.from_user.id][2] > 20 or \
            (str(s.cell(row=users[message.from_user.id][0] + 2, column=5).value) == str(2) and
             users[message.from_user.id][2] > 10):
        await message.answer('Введите ответ на 1 вопрос')
    else:
        await message.answer('Введите ответ на ' + str(users[message.from_user.id][2]) +
                             ' вопрос')
    state = dp.current_state(user=message.from_user.id)
    await state.set_state(TestStates.all()[3])


@dp.message_handler(state=TestStates.TEST_STATE_3)
async def process_ans_command(message: types.Message):
    w = openpyxl.load_workbook('users.xlsx')
    s = w['users']
    s = w.active

    if str(s.cell(row=users[message.from_user.id][0] + 2, column=5).value) == str(1) and users[message.from_user.id][2] \
            <= 20:
        sh = sh1
    elif (str(s.cell(row=users[message.from_user.id][0] + 2, column=5).value) == str(1) and
          users[message.from_user.id][2] > 20) \
            or (str(s.cell(row=users[message.from_user.id][0] + 2, column=5).value) == str(2) and
                users[message.from_user.id][2] <= 10):
        sh = sh2
        cell = s.cell(row=users[message.from_user.id][0] + 2, column=5)
        cell.value = 2
        w.save('users.xlsx')
        w.close()
        if users[message.from_user.id][2] > 10:
            users[message.from_user.id][2] = 1
    else:
        sh = sh3
        cell = s.cell(row=users[message.from_user.id][0] + 2, column=5)
        cell.value = 3
        w.save('users.xlsx')
        w.close()
        if users[message.from_user.id][2] > 10:
            users[message.from_user.id][2] = 1

    if (message.text == str(sh.cell(row=users[message.from_user.id][2], column=1).value)) or (
            message.text == str(sh.cell(row=users[message.from_user.id][2], column=2).value)) or (
            message.text == str(sh.cell(row=users[message.from_user.id][2], column=3).value)) or (
            message.text == str(sh.cell(row=users[message.from_user.id][2], column=4).value)):
        if ((users[message.from_user.id][2] <= 10 and str(
                s.cell(row=users[message.from_user.id][0] + 2, column=5).value) == str(1)) or (
                users[message.from_user.id][2] <= 5)):
            users[message.from_user.id][1] = users[message.from_user.id][1] + 1
        else:
            users[message.from_user.id][1] = users[message.from_user.id][1] + 2
    state = dp.current_state(user=message.from_user.id)
    if users[message.from_user.id][2] == 20 and str(
            s.cell(row=users[message.from_user.id][0] + 2, column=5).value) == str(1):
        if message.from_user.id == 765839138:
            await state.set_state(TestStates.all()[0])
        else:
            await state.set_state(TestStates.all()[2])
        await message.answer(MESSAGES['first_e'])
        users[message.from_user.id][2] += 1
        save_s(message)

    elif users[message.from_user.id][2] == 10 and str(
            s.cell(row=users[message.from_user.id][0] + 2, column=5).value) == str(2):
        await message.answer(MESSAGES['second_m'])
        save_s(message)

        users[message.from_user.id][2] += 1
        if message.from_user.id == 765839138:
            await state.set_state(TestStates.all()[0])
        else:
            await state.set_state(TestStates.all()[2])
        await process_quiz_command(message)
    elif users[message.from_user.id][2] == 10 and str(
            s.cell(row=users[message.from_user.id][0] + 2, column=5).value) == str(3):
        await message.answer(MESSAGES['second_e'])
        if message.from_user.id == 765839138:
            await state.set_state(TestStates.all()[0])
        else:
            await state.set_state(TestStates.all()[5])
        save_s(message)
    else:
        save_s(message)
        users[message.from_user.id][2] = users[message.from_user.id][2] + 1
        await state.set_state(TestStates.all()[2])
        await process_quiz_command(message)


@dp.message_handler(state=TestStates.TEST_STATE_6)
async def process_fio_command(message: types.Message):
    wb = openpyxl.load_workbook('users.xlsx')
    sheet = wb['users']
    sheet = wb.active
    cell = sheet.cell(row=users[message.from_user.id][0] + 2, column=2)
    cell.value = message.text
    wb.save('users.xlsx')
    wb.close()

    state = dp.current_state(user=message.from_user.id)
    await state.set_state(TestStates.all()[7])
    await message.answer('Прикрепи, пожалуйста, ссылку на ВК, чтобы мы могли связаться с тобой :)')


@dp.message_handler(state=TestStates.TEST_STATE_7)
async def process_link_command(message: types.Message):
    wb = openpyxl.load_workbook('users.xlsx')
    sheet = wb['users']
    sheet = wb.active
    cell = sheet.cell(row=users[message.from_user.id][0] + 2, column=4)
    cell.value = message.text
    wb.save('users.xlsx')
    wb.close()

    state = dp.current_state(user=message.from_user.id)
    if message.from_user.id == 765839138:
        await state.set_state(TestStates.all()[0])
    else:
        await state.set_state(TestStates.all()[2])
    await message.answer('Регистрация успешно пройдена!'
                         '\n\nНажми —> /info <— чтобы узнать, что делать дальше')


@dp.message_handler(state=TestStates.TEST_STATE_0 | TestStates.TEST_STATE_5, commands=['score'])
async def process_score_command(message: types.Message):
    await message.answer('Твой счет: ' + str(users[message.from_user.id][1]) + ' баллов из 60'
                                                                               '\nЛидеры будут объявлены позже')


@dp.message_handler(state=TestStates.TEST_STATE_0, commands=['res'])
async def process_score_command(message: types.Message):
    f = open("users.xlsx", "rb")
    await message.answer_document(f)


@dp.message_handler(state=TestStates.all())
async def process_no_command(message: types.Message):
    await message.answer(MESSAGES['no_command'])


@dp.message_handler()
async def echo_message(msg: types.Message):
    await bot.send_message(msg.from_user.id, '/start чтобы начать')


async def shutdown(dispatcher: Dispatcher):
    await dispatcher.storage.close()
    await dispatcher.storage.wait_closed()


if __name__ == '__main__':
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    button_1 = types.KeyboardButton(
        text="/start")  # тут мы делаем кнопку старт, если нужно добавить другие или редактировать эту - нужно посмотреть руководство
    keyboard.add(
        button_1)  # https://mastergroosha.github.io/telegram-tutorial-2/buttons/  понятное руководство по кнопкам
    executor.start_polling(dp, on_shutdown=shutdown)
